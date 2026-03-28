from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date
from io import BytesIO
import hashlib
from pathlib import Path
from typing import Callable, List, Optional, Tuple
import logging
import re
import traceback
import uuid

import pandas as pd
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from sqlalchemy import Index, inspect, text

db = SQLAlchemy()

logger = logging.getLogger(__name__)
BALANCE_TOLERANCE = 0.005

# ==================== CATÁLOGO DE ENTES (hardcoded) ====================

ENTES_CATALOG = [
    # ── Poder Ejecutivo y Dependencias (num 1 – 1.29) ──
    {"codigo": "1",    "siglas": "EJECUTIVO",   "nombre": "PODER EJECUTIVO DEL ESTADO DE TLAXCALA",                                  "tipo": "PODER DEL ESTADO",                      "ambito": "ESTATAL"},
    {"codigo": "1.1",  "siglas": "DG",          "nombre": "DESPACHO DE LA GOBERNADORA",                                              "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.2",  "siglas": "SEGOB",       "nombre": "SECRETARÍA DE GOBIERNO",                                                  "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.3",  "siglas": "OMG",         "nombre": "OFICIALIA MAYOR DE GOBIERNO",                                             "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.4",  "siglas": "SF",          "nombre": "SECRETARÍA DE FINANZAS",                                                  "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.5",  "siglas": "SEDECO",      "nombre": "SECRETARÍA DE DESARROLLO ECONOMICO",                                      "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.6",  "siglas": "SECTUR",      "nombre": "SECRETARÍA DE TURISMO",                                                   "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.7",  "siglas": "SI",          "nombre": "SECRETARÍA DE INFRAESTRUCTURA",                                           "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.8",  "siglas": "SEPE",        "nombre": "SECRETARÍA DE EDUCACIÓN PÚBLICA",                                         "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.9",  "siglas": "SMYT",        "nombre": "SECRETARÍA DE MOVILIDAD Y TRANSPORTE",                                    "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.10", "siglas": "SFP",         "nombre": "SECRETARÍA DE LA FUNCIÓN PÚBLICA",                                        "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.11", "siglas": "SIA",         "nombre": "SECRETARÍA DE IMPULSO AGROPECUARIO",                                      "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.12", "siglas": "CCOM",        "nombre": "COORDINACIÓN DE COMUNICACIÓN",                                            "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.13", "siglas": "SMA",         "nombre": "SECRETARÍA DE MEDIO AMBIENTE",                                            "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.14", "siglas": "SC",          "nombre": "SECRETARÍA DE CULTURA",                                                   "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.15", "siglas": "SMET",        "nombre": "SECRETARÍA DE LAS MUJERES",                                               "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.16", "siglas": "SOTYV",       "nombre": "SECRETARÍA DE ORDENAMIENTO TERRITORIAL Y VIVIENDA",                       "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.17", "siglas": "SSC",         "nombre": "SECRETARÍA DE SEGURIDAD CIUDADANA",                                       "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.18", "siglas": "CGPI",        "nombre": "COORDINACIÓN GENERAL DE PLANEACIÓN E INVERSIÓN",                          "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.19", "siglas": "SB",          "nombre": "SECRETARÍA DE BIENESTAR",                                                 "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.20", "siglas": "STYC",        "nombre": "SECRETARÍA DE TRABAJO Y COMPETITIVIDAD",                                  "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.21", "siglas": "CJE",         "nombre": "CONSEJERÍA JURÍDICA DEL EJECUTIVO",                                       "tipo": "DEPENDENCIA",                           "ambito": "ESTATAL"},
    {"codigo": "1.22", "siglas": "CEPC",        "nombre": "COORDINACIÓN ESTATAL DE PROTECCIÓN CIVIL",                                "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    {"codigo": "1.23", "siglas": "SESESP",      "nombre": "SECRETARIADO EJECUTIVO DEL SISTEMA ESTATAL DE SEGURIDAD PÚBLICA",         "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    {"codigo": "1.24", "siglas": "ITDT",        "nombre": "INSTITUTO TLAXCALTECA DE DESARROLLO TAURINO",                             "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    {"codigo": "1.25", "siglas": "ITAES",       "nombre": "INSTITUTO TLAXCALTECA DE ASISTENCIA ESPECIALIZADA A LA SALUD",            "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    {"codigo": "1.26", "siglas": "CEAM",        "nombre": "COMISIÓN ESTATAL DE ARBITRAJE MÉDICO",                                    "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    {"codigo": "1.27", "siglas": "CAT",         "nombre": "CASA DE LAS ARTESANÍAS DE TLAXCALA",                                      "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    {"codigo": "1.28", "siglas": "PROPAET",     "nombre": "PROCURADURÍA DE PROTECCIÓN AL AMBIENTE DEL ESTADO DE TLAXCALA",           "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    {"codigo": "1.29", "siglas": "IFAST",       "nombre": "INSTITUTO DE FAUNA SILVESTRE PARA EL ESTADO DE TLAXCALA",                 "tipo": "DESCONCENTRADOS",                       "ambito": "ESTATAL"},
    # ── Poderes del Estado (num 2, 3) ──
    {"codigo": "2",    "siglas": "LEGISLATIVO",  "nombre": "PODER LEGISLATIVO DEL ESTADO DE TLAXCALA",                               "tipo": "PODERES DEL ESTADO",                    "ambito": "ESTATAL"},
    {"codigo": "3",    "siglas": "PJET",         "nombre": "PODER JUDICIAL DEL ESTADO DE TLAXCALA",                                  "tipo": "PODERES DEL ESTADO",                    "ambito": "ESTATAL"},
    # ── Descentralizados / Paraestatales (num 4 – 32) ──
    {"codigo": "4",    "siglas": "CCLET",        "nombre": "CENTRO DE CONCILIACIÓN LABORAL DEL ESTADO DE TLAXCALA",                  "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "5",    "siglas": "CEAS",         "nombre": "COMISIÓN ESTATAL DEL AGUA Y SANEAMIENTO DEL ESTADO DE TLAXCALA",         "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "6",    "siglas": "COBAT",        "nombre": "COLEGIO DE BACHILLERES DEL ESTADO DE TLAXCALA",                          "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "7",    "siglas": "CONALEP",      "nombre": "COLEGIO DE EDUCACIÓN PROFESIONAL TÉCNICA DEL ESTADO DE TLAXCALA",        "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "8",    "siglas": "CECYTE",       "nombre": "COLEGIO DE ESTUDIOS CIENTÍFICOS Y TECNOLÓGICOS DEL ESTADO DE TLAXCALA",  "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "9",    "siglas": "COESPO",       "nombre": "CONSEJO ESTATAL DE POBLACIÓN",                                           "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "10",   "siglas": "CORACYT",      "nombre": "COORDINACIÓN DE RADIO, CINE Y TELEVISIÓN",                               "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "11",   "siglas": "COLTLAX",      "nombre": "EL COLEGIO DE TLAXCALA, A.C.",                                           "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "12",   "siglas": "FIDECIX",      "nombre": "FIDEICOMISO DE LA CIUDAD INDUSTRIAL DE XICOTÉNCATL",                     "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "13",   "siglas": "CEAVIT",       "nombre": "COMISIÓN EJECUTIVA DE ATENCIÓN A VÍCTIMAS DEL ESTADO DE TLAXCALA",       "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "14",   "siglas": "FOMTLAX",      "nombre": "FONDO MACRO PARA EL DESARROLLO INTEGRAL DE TLAXCALA",                   "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "15",   "siglas": "ICATLAX",      "nombre": "INSTITUTO DE CAPACITACIÓN PARA EL TRABAJO DEL ESTADO DE TLAXCALA",       "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "16",   "siglas": "IDC",          "nombre": "INSTITUTO DE CATASTRO DEL ESTADO DE TLAXCALA",                           "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "17",   "siglas": "IDET",         "nombre": "INSTITUTO DEL DEPORTE DE TLAXCALA",                                      "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "18",   "siglas": "ITST",         "nombre": "INSTITUTO TECNOLÓGICO SUPERIOR DE TLAXCO",                               "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "19",   "siglas": "ITIFE",        "nombre": "INSTITUTO TLAXCALTECA DE LA INFRAESTRUCTURA FÍSICA EDUCATIVA",           "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "20",   "siglas": "ITJ",          "nombre": "INSTITUTO TLAXCALTECA DE LA JUVENTUD",                                   "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "21",   "siglas": "ITEA",         "nombre": "INSTITUTO TLAXCALTECA PARA LA EDUCACIÓN DE LOS ADULTOS, ITEA",           "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "22",   "siglas": "OPD_SALUD",    "nombre": "ÓRGANISMO PÚBLICO DESCENTRALIZADO SALUD DE TLAXCALA",                    "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "23",   "siglas": "CRI-ESCUELA",  "nombre": "PATRONATO CENTRO DE REHABILITACIÓN INTEGRAL Y ESCUELA EN TERAPIA FÍSICA Y REHABILITACIÓN", "tipo": "DESCENTRALIZADO/PARAESTATAL", "ambito": "ESTATAL"},
    {"codigo": "24",   "siglas": "LA_LIBERTAD",  "nombre": "PATRONATO \"LA LIBERTAD CENTRO CULTURAL DE APIZACO\"",                   "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "25",   "siglas": "PCET",         "nombre": "PENSIONES CIVILES DEL ESTADO DE TLAXCALA",                               "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "26",   "siglas": "SEDIF",        "nombre": "SISTEMA ESTATAL PARA EL DESARROLLO INTEGRAL DE LA FAMILIA",              "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "27",   "siglas": "USET",         "nombre": "UNIDAD DE SERVICIOS EDUCATIVOS DEL ESTADO DE TLAXCALA",                  "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "28",   "siglas": "UPT",          "nombre": "UNIVERSIDAD POLITÉCNICA DE TLAXCALA",                                    "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "29",   "siglas": "UTREP",        "nombre": "UNIVERSIDAD POLITÉCNICA DE TLAXCALA REGIÓN PONIENTE",                    "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "30",   "siglas": "UTT",          "nombre": "UNIVERSIDAD TECNOLÓGICA DE TLAXCALA",                                    "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "31",   "siglas": "UIT",          "nombre": "UNIVERSIDAD INTERCULTURAL DE TLAXCALA",                                  "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "32",   "siglas": "AGHET",        "nombre": "ARCHIVO GENERAL E HISTORICO DEL ESTADO DE TLAXCALA",                     "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    # ── Organismos Autónomos (num 33 – 42) ──
    {"codigo": "33",   "siglas": "TJA",          "nombre": "TRIBUNAL DE JUSTICIA ADMINISTRATIVA DEL ESTADO DE TLAXCALA",             "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "34",   "siglas": "UAT",          "nombre": "UNIVERSIDAD AUTÓNOMA DE TLAXCALA",                                       "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "35",   "siglas": "CEDH",         "nombre": "COMISIÓN ESTATAL DE DERECHOS HUMANOS",                                   "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "36",   "siglas": "ITE",          "nombre": "INSTITUTO TLAXCALTECA DE ELECCIONES",                                    "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "37",   "siglas": "IAIP",         "nombre": "INSTITUTO DE ACCESO A LA INFORMACIÓN PÚBLICA Y PROTECCIÓN DE DATOS PERSONALES DEL ESTADO DE TLAXCALA", "tipo": "ORGANISMO AUTÓNOMO", "ambito": "ESTATAL"},
    {"codigo": "38",   "siglas": "TCYA",         "nombre": "TRIBUNAL DE CONCILIACIÓN Y ARBITRAJE DEL ESTADO DE TLAXCALA",            "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "39",   "siglas": "TET",          "nombre": "TRIBUNAL ELECTORAL DE TLAXCALA",                                         "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "40",   "siglas": "FGJET",        "nombre": "FISCALÍA GENERAL DE JUSTICIA DEL ESTADO DE TLAXCALA",                    "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "41",   "siglas": "SESAET",       "nombre": "SECRETARIA EJECUTIVA DEL SISTEMA ANTICORRUPCIÓN DEL ESTADO DE TLAXCALA", "tipo": "ORGANISMO DESCENTRALIZADO NO SECTORIZADO", "ambito": "ESTATAL"},
    {"codigo": "42",   "siglas": "P_FERIA",      "nombre": "PATRONATO PARA LAS EXPOSICIONES Y FERIAS EN LA CIUDAD DE TLAXCALA",      "tipo": "PATRONATO",                             "ambito": "ESTATAL"},
    # ── Municipios (60) ──
    {"codigo": "M01",  "siglas": "ACUAMANALA",    "nombre": "ACUAMANALA DE MIGUEL HIDALGO",                                          "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M02",  "siglas": "ATLTZAYANCA",   "nombre": "ATLTZAYANCA",                                                           "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M03",  "siglas": "AMAXAC",        "nombre": "AMAXAC DE GUERRERO",                                                    "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M04",  "siglas": "APETATITLAN",   "nombre": "APETATITLÁN DE ANTONIO CARVAJAL",                                       "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M05",  "siglas": "APIZACO",       "nombre": "APIZACO",                                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M06",  "siglas": "ATLANGATEPEC",  "nombre": "ATLANGATEPEC",                                                          "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M07",  "siglas": "BENITO JUÁREZ", "nombre": "BENITO JUÁREZ",                                                         "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M08",  "siglas": "CALPULALPAN",   "nombre": "CALPULALPAN",                                                           "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M09",  "siglas": "CHIAUTEMPAN",   "nombre": "CHIAUTEMPAN",                                                           "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M10",  "siglas": "CONTLA",        "nombre": "CONTLA DE JUAN CUAMATZI",                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M11",  "siglas": "CUAPIAXTLA",    "nombre": "CUAPIAXTLA",                                                            "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M12",  "siglas": "CUAXOMULCO",    "nombre": "CUAXOMULCO",                                                            "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M13",  "siglas": "EL CARMEN",     "nombre": "EL CARMEN TEQUEXQUITLA",                                                "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M14",  "siglas": "EMILIANO Z",    "nombre": "EMILIANO ZAPATA",                                                       "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M15",  "siglas": "ESPAÑITA",      "nombre": "ESPAÑITA",                                                              "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M16",  "siglas": "HUAMANTLA",     "nombre": "HUAMANTLA",                                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M17",  "siglas": "HUEYOTLIPAN",   "nombre": "HUEYOTLIPAN",                                                           "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M18",  "siglas": "IXTACUIXTLA",   "nombre": "IXTACUIXTLA DE MARIANO MATAMOROS",                                      "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M19",  "siglas": "IXTENCO",       "nombre": "IXTENCO",                                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M20",  "siglas": "LA MAGDALENA",  "nombre": "LA MAGDALENA TLALTELULCO",                                              "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M21",  "siglas": "L. CARDENAS",   "nombre": "LÁZARO CÁRDENAS",                                                       "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M22",  "siglas": "MAZATECOCHCO",  "nombre": "MAZATECOCHCO DE JOSÉ MARÍA MORELOS",                                    "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M23",  "siglas": "MUÑOZ",         "nombre": "MUÑOZ DE DOMINGO ARENAS",                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M24",  "siglas": "NANACAMILPA",   "nombre": "NANACAMILPA DE MARIANO ARISTA",                                         "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M25",  "siglas": "NATIVITAS",     "nombre": "NATIVITAS",                                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M26",  "siglas": "PANOTLA",       "nombre": "PANOTLA",                                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M27",  "siglas": "PAPALOTLA",     "nombre": "PAPALOTLA DE XICOHTÉNCATL",                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M28",  "siglas": "TEXOLOC",       "nombre": "SAN DAMIÁN TEXOLOC",                                                    "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M29",  "siglas": "TETLANOHCAN",   "nombre": "SAN FRANCISCO TETLANOHCAN",                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M30",  "siglas": "ZACUALPAN",     "nombre": "SAN JERÓNIMO ZACUALPAN",                                                "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M31",  "siglas": "S.J. TEACALCO", "nombre": "SAN JOSÉ TEACALCO",                                                     "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M32",  "siglas": "HUACTZINCO",    "nombre": "SAN JUAN HUACTZINCO",                                                   "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M33",  "siglas": "AXOCOMANITLA",  "nombre": "SAN LORENZO AXOCOMANITLA",                                              "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M34",  "siglas": "TECOPILCO",     "nombre": "SAN LUCAS TECOPILCO",                                                   "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M35",  "siglas": "SN. PABLO",     "nombre": "SAN PABLO DEL MONTE",                                                   "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M36",  "siglas": "S.L. CARDENAS", "nombre": "SANCTÓRUM DE LÁZARO CÁRDENAS",                                          "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M37",  "siglas": "NOPALUCAN",     "nombre": "SANTA ANA NOPALUCAN",                                                   "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M38",  "siglas": "S.A. TEACALCO", "nombre": "SANTA APOLONIA TEACALCO",                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M39",  "siglas": "AYOMETLA",      "nombre": "SANTA CATARINA AYOMETLA",                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M40",  "siglas": "QUILEHTLA",     "nombre": "SANTA CRUZ QUILEHTLA",                                                  "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M41",  "siglas": "S.C. TLAXCALA", "nombre": "SANTA CRUZ TLAXCALA",                                                   "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M42",  "siglas": "XILOXOXTLA",    "nombre": "SANTA ISABEL XILOXOXTLA",                                              "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M43",  "siglas": "TENANCINGO",    "nombre": "TENANCINGO",                                                            "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M44",  "siglas": "TEOLOCHOLCO",   "nombre": "TEOLOCHOLCO",                                                           "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M45",  "siglas": "TEPETITLA",     "nombre": "TEPETITLA DE LARDIZÁBAL",                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M46",  "siglas": "TEPEYANCO",     "nombre": "TEPEYANCO",                                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M47",  "siglas": "TERRENATE",     "nombre": "TERRENATE",                                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M48",  "siglas": "T. SOLIDARIDAD", "nombre": "TETLA DE LA SOLIDARIDAD",                                              "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M49",  "siglas": "TETLATLAHUCA",  "nombre": "TETLATLAHUCA",                                                          "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M50",  "siglas": "TLAXCALA",      "nombre": "TLAXCALA",                                                              "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M51",  "siglas": "TLAXCO",        "nombre": "TLAXCO",                                                                "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M52",  "siglas": "TOCATLÁN",      "nombre": "TOCATLÁN",                                                              "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M53",  "siglas": "TOTOLAC",       "nombre": "TOTOLAC",                                                               "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M54",  "siglas": "TZOMPANTEPEC",  "nombre": "TZOMPANTEPEC",                                                          "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M55",  "siglas": "XALOZTOC",      "nombre": "XALOZTOC",                                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M56",  "siglas": "XALTOCAN",      "nombre": "XALTOCAN",                                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M57",  "siglas": "XICOHTZINCO",   "nombre": "XICOHTZINCO",                                                          "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M58",  "siglas": "YAUHQUEMEHCAN", "nombre": "YAUHQUEMEHCAN",                                                         "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M59",  "siglas": "ZACATELCO",     "nombre": "ZACATELCO",                                                             "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    {"codigo": "M60",  "siglas": "ZITLALTEPEC",   "nombre": "ZITLALTÉPEC DE TRINIDAD SÁNCHEZ SANTOS",                                "tipo": "MUNICIPIO",                             "ambito": "MUNICIPAL"},
    # ── Paramunicipales (5) ──
    {"codigo": "148",  "siglas": "CAPAMH",        "nombre": "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE HUAMANTLA",   "tipo": "PARAMUNICIPAL",                         "ambito": "MUNICIPAL"},
    {"codigo": "150",  "siglas": "CAPAMA",        "nombre": "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE APIZACO",     "tipo": "PARAMUNICIPAL",                         "ambito": "MUNICIPAL"},
    {"codigo": "151",  "siglas": "CAPACH",        "nombre": "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE CHIAUTEMPAN", "tipo": "PARAMUNICIPAL",                         "ambito": "MUNICIPAL"},
    {"codigo": "154",  "siglas": "CAPAZ",         "nombre": "COMISIÓN DE AGUA POTABLE Y ALCANTARILLADO DEL MUNICIPIO DE ZACATELCO",   "tipo": "PARAMUNICIPAL",                         "ambito": "MUNICIPAL"},
    {"codigo": "156",  "siglas": "CAPAMT",        "nombre": "COMISIÓN DE POTABLE Y Y ALCANTARILLADO DEL MUNICIPIO TLAXCALA",          "tipo": "PARAMUNICIPAL",                         "ambito": "MUNICIPAL"},
    # ── Legacy SAACG codes (backwards compatibility) ──
    {"codigo": "101",  "siglas": "ITE",          "nombre": "INSTITUTO TLAXCALTECA DE ELECCIONES",                                     "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "102",  "siglas": "UAT",          "nombre": "UNIVERSIDAD AUTÓNOMA DE TLAXCALA",                                        "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "103",  "siglas": "CEDH",         "nombre": "COMISIÓN ESTATAL DE DERECHOS HUMANOS",                                    "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "104",  "siglas": "TCYA",         "nombre": "TRIBUNAL DE CONCILIACIÓN Y ARBITRAJE DEL ESTADO DE TLAXCALA",             "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "105",  "siglas": "IAIPTLAX",     "nombre": "INSTITUTO DE ACCESO A LA INFORMACIÓN PÚBLICA Y PROTECCIÓN DE DATOS PERSONALES DEL ESTADO DE TLAXCALA", "tipo": "ORGANISMO AUTÓNOMO", "ambito": "ESTATAL"},
    {"codigo": "106",  "siglas": "PODER JUDICIAL", "nombre": "PODER JUDICIAL DEL ESTADO DE TLAXCALA",                                "tipo": "PODERES DEL ESTADO",                    "ambito": "ESTATAL"},
    {"codigo": "107",  "siglas": "CONGRESO",     "nombre": "PODER LEGISLATIVO DEL ESTADO DE TLAXCALA",                                "tipo": "PODERES DEL ESTADO",                    "ambito": "ESTATAL"},
    {"codigo": "110",  "siglas": "CONALEP",      "nombre": "COLEGIO DE EDUCACIÓN PROFESIONAL TÉCNICA DEL ESTADO DE TLAXCALA",         "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "116",  "siglas": "CORACYT",      "nombre": "COORDINACIÓN DE RADIO, CINE Y TELEVISIÓN",                                "tipo": "DESCENTRALIZADO/PARAESTATAL",           "ambito": "ESTATAL"},
    {"codigo": "149",  "siglas": "TET",          "nombre": "TRIBUNAL ELECTORAL DE TLAXCALA",                                          "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "152",  "siglas": "TJA",          "nombre": "TRIBUNAL DE JUSTICIA ADMINISTRATIVA DEL ESTADO DE TLAXCALA",              "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
    {"codigo": "153",  "siglas": "SESAET",       "nombre": "SECRETARÍA EJECUTIVA DEL SISTEMA ANTICORRUPCIÓN DEL ESTADO DE TLAXCALA",  "tipo": "ORGANISMO DESCENTRALIZADO NO SECTORIZADO", "ambito": "ESTATAL"},
    {"codigo": "157",  "siglas": "FISCALIA",     "nombre": "FISCALÍA GENERAL DE JUSTICIA DEL ESTADO DE TLAXCALA",                     "tipo": "ORGANISMO AUTÓNOMO",                    "ambito": "ESTATAL"},
]

ENTES_BY_CODIGO = {e["codigo"]: e for e in ENTES_CATALOG}


# ==================== MODELOS ====================

class Transaccion(db.Model):
    __tablename__ = "transacciones"

    id = db.Column(db.Integer, primary_key=True)

    # Carga
    lote_id = db.Column(db.String(36), nullable=False, index=True)
    archivo_origen = db.Column(db.String(255), nullable=False)
    fecha_carga = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    usuario_carga = db.Column(db.String(100))

    # Cuenta contable
    cuenta_contable = db.Column(db.String(64), nullable=False, index=True)
    nombre_cuenta = db.Column(db.Text)

    # Segmentos de cuenta (separados por punto en TXT)
    seg1 = db.Column(db.String(10), index=True)
    seg2 = db.Column(db.String(10), index=True)
    seg3 = db.Column(db.String(10), index=True)
    seg4 = db.Column(db.String(10), index=True)
    seg5 = db.Column(db.String(10), index=True)
    seg6 = db.Column(db.String(10), index=True)

    # Componentes de cuenta (formato vertical SIIF)
    genero = db.Column(db.String(1), index=True)
    grupo = db.Column(db.String(1), index=True)
    rubro = db.Column(db.String(1), index=True)
    cuenta_nivel = db.Column(db.String(1), index=True)
    subcuenta = db.Column(db.String(1), index=True)
    dependencia = db.Column(db.String(2), index=True)
    unidad_responsable = db.Column(db.String(2), index=True)
    centro_costo = db.Column(db.String(2), index=True)
    proyecto_presupuestario = db.Column(db.String(2), index=True)
    fuente = db.Column(db.String(1), index=True)
    subfuente = db.Column(db.String(2), index=True)
    tipo_recurso = db.Column(db.String(1), index=True)
    partida_presupuestal = db.Column(db.String(4), index=True)

    # Ente
    ente_codigo = db.Column(db.String(20), index=True)
    ente_nombre = db.Column(db.String(255), index=True)
    ente_siglas = db.Column(db.String(50))

    # Descripciones expandidas (líneas de descripción del TXT)
    descripcion_1 = db.Column(db.Text)
    descripcion_2 = db.Column(db.Text)
    descripcion_3 = db.Column(db.Text)

    # Transacción
    fecha_transaccion = db.Column(db.Date, nullable=False, index=True)
    poliza = db.Column(db.String(50), index=True)
    primer_concepto = db.Column(db.Text)
    segundo_concepto = db.Column(db.Text)
    tercer_concepto = db.Column(db.Text)
    concepto_especifico = db.Column(db.Text)
    concepto = db.Column(db.Text)
    requisicion = db.Column(db.String(100))
    cheque_folio = db.Column(db.String(100))
    beneficiario_codigo = db.Column(db.String(20))
    beneficiario = db.Column(db.Text)
    descripcion = db.Column(db.Text)

    # Obra
    obra_accion = db.Column(db.String(100))
    nombre_obra = db.Column(db.Text)

    # Concepto presupuestal
    concept_codigo = db.Column(db.String(20))
    nombre_concepto = db.Column(db.Text)

    # Fuente de financiamiento
    ff = db.Column(db.String(10))
    rem = db.Column(db.String(10))
    nombre_fuente_financiamiento = db.Column(db.Text)

    # Unidad administrativa
    unidad_codigo = db.Column(db.String(10))
    nombre_unidad_administrativa = db.Column(db.Text)

    # Programa
    programa = db.Column(db.String(50))
    nombre_programa = db.Column(db.Text)

    # Componente
    componente = db.Column(db.String(10))
    nombre_componente = db.Column(db.Text)

    # Actividad
    actividad = db.Column(db.String(10))
    nombre_actividad = db.Column(db.Text)

    # Función
    funcion = db.Column(db.String(10))
    nombre_funcion = db.Column(db.Text)

    # Otros
    contrato = db.Column(db.String(100))
    ficha = db.Column(db.String(100))
    referencia = db.Column(db.Text)
    persona = db.Column(db.Text)
    rfc = db.Column(db.String(50))
    nombre_persona = db.Column(db.Text)
    numero_bien = db.Column(db.String(100))
    nombre_bien = db.Column(db.Text)
    unidades = db.Column(db.String(50))

    # Partida presupuestal (texto completo del TXT)
    partida_presupuestal_txt = db.Column(db.String(100))
    nombre_partida_presupuestal = db.Column(db.Text)

    # Región y Municipio
    region = db.Column(db.String(10))
    nombre_region = db.Column(db.Text)
    municipio = db.Column(db.String(10))
    nombre_municipio = db.Column(db.Text)

    # Montos
    cargos = db.Column(db.Numeric(15, 2), default=0)
    abonos = db.Column(db.Numeric(15, 2), default=0)
    saldo_final = db.Column(db.Numeric(15, 2), default=0)

    # Montos finales (últimas 3 columnas del TXT)
    cargos_f = db.Column(db.Numeric(15, 2), default=0)
    abonos_f = db.Column(db.Numeric(15, 2), default=0)
    saldo_f = db.Column(db.Numeric(15, 2), default=0)

    hash_registro = db.Column(db.String(64), unique=True, index=True)

    __table_args__ = (
        Index("idx_cuenta_fecha", "cuenta_contable", "fecha_transaccion"),
        Index("idx_ente_fecha", "ente_codigo", "fecha_transaccion"),
        Index("idx_lote_cuenta", "lote_id", "cuenta_contable"),
    )

    def to_dict(self):
        return {
            "id": self.id,
            "lote_id": self.lote_id,
            "archivo_origen": self.archivo_origen,
            "fecha_carga": self.fecha_carga.isoformat() if self.fecha_carga else None,
            "cuenta_contable": self.cuenta_contable,
            "nombre_cuenta": self.nombre_cuenta,
            "descripcion_1": self.descripcion_1 or "",
            "descripcion_2": self.descripcion_2 or "",
            "descripcion_3": self.descripcion_3 or "",
            "seg1": self.seg1,
            "seg2": self.seg2,
            "seg3": self.seg3,
            "seg4": self.seg4,
            "seg5": self.seg5,
            "seg6": self.seg6,
            "genero": self.genero,
            "grupo": self.grupo,
            "rubro": self.rubro,
            "cuenta_nivel": self.cuenta_nivel,
            "subcuenta": self.subcuenta,
            "dependencia": self.dependencia,
            "unidad_responsable": self.unidad_responsable,
            "centro_costo": self.centro_costo,
            "proyecto_presupuestario": self.proyecto_presupuestario,
            "fuente": self.fuente,
            "subfuente": self.subfuente,
            "tipo_recurso": self.tipo_recurso,
            "partida_presupuestal": self.partida_presupuestal,
            "ente_codigo": self.ente_codigo,
            "ente_nombre": self.ente_nombre,
            "ente_siglas": self.ente_siglas,
            "fecha_transaccion": self.fecha_transaccion.strftime("%d/%m/%Y") if self.fecha_transaccion else None,
            "poliza": self.poliza,
            "primer_concepto": self.primer_concepto or "",
            "segundo_concepto": self.segundo_concepto or "",
            "tercer_concepto": self.tercer_concepto or "",
            "concepto_especifico": self.concepto_especifico or "",
            "concepto": self.concepto or "",
            "requisicion": self.requisicion or "",
            "cheque_folio": self.cheque_folio,
            "beneficiario_codigo": self.beneficiario_codigo or "",
            "beneficiario": self.beneficiario,
            "descripcion": self.descripcion,
            "obra_accion": self.obra_accion or "",
            "nombre_obra": self.nombre_obra or "",
            "concept_codigo": self.concept_codigo or "",
            "nombre_concepto": self.nombre_concepto or "",
            "ff": self.ff or "",
            "rem": self.rem or "",
            "nombre_fuente_financiamiento": self.nombre_fuente_financiamiento or "",
            "unidad_codigo": self.unidad_codigo or "",
            "nombre_unidad_administrativa": self.nombre_unidad_administrativa or "",
            "programa": self.programa or "",
            "nombre_programa": self.nombre_programa or "",
            "componente": self.componente or "",
            "nombre_componente": self.nombre_componente or "",
            "actividad": self.actividad or "",
            "nombre_actividad": self.nombre_actividad or "",
            "funcion": self.funcion or "",
            "nombre_funcion": self.nombre_funcion or "",
            "contrato": self.contrato or "",
            "ficha": self.ficha or "",
            "referencia": self.referencia or "",
            "persona": self.persona or "",
            "rfc": self.rfc or "",
            "nombre_persona": self.nombre_persona or "",
            "numero_bien": self.numero_bien or "",
            "nombre_bien": self.nombre_bien or "",
            "unidades": self.unidades or "",
            "partida_presupuestal_txt": self.partida_presupuestal_txt or "",
            "nombre_partida_presupuestal": self.nombre_partida_presupuestal or "",
            "region": self.region or "",
            "nombre_region": self.nombre_region or "",
            "municipio": self.municipio or "",
            "nombre_municipio": self.nombre_municipio or "",
            "cargos": float(self.cargos) if self.cargos else 0,
            "abonos": float(self.abonos) if self.abonos else 0,
            "saldo_final": float(self.saldo_final) if self.saldo_final else 0,
            "cargos_f": float(self.cargos_f) if self.cargos_f else 0,
            "abonos_f": float(self.abonos_f) if self.abonos_f else 0,
            "saldo_f": float(self.saldo_f) if self.saldo_f else 0,
        }


class LoteCarga(db.Model):
    __tablename__ = "lotes_carga"

    id = db.Column(db.Integer, primary_key=True)
    lote_id = db.Column(db.String(36), unique=True, nullable=False, index=True)
    fecha_carga = db.Column(db.DateTime, default=datetime.utcnow)
    usuario = db.Column(db.String(100))
    archivos = db.Column(db.JSON)
    total_registros = db.Column(db.Integer, default=0)
    estado = db.Column(db.String(20), default="procesando")
    mensaje = db.Column(db.Text)

    def to_dict(self):
        return {
            "id": self.id,
            "lote_id": self.lote_id,
            "fecha_carga": self.fecha_carga.isoformat() if self.fecha_carga else None,
            "usuario": self.usuario,
            "archivos": self.archivos,
            "total_registros": self.total_registros,
            "estado": self.estado,
            "mensaje": self.mensaje,
        }


class CargaJob(db.Model):
    __tablename__ = "carga_jobs"

    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.String(36), unique=True, nullable=False, index=True)
    usuario = db.Column(db.String(100))
    ente_nombre = db.Column(db.String(255))
    archivos = db.Column(db.JSON)
    progress = db.Column(db.Integer, default=0)
    message = db.Column(db.Text)
    done = db.Column(db.Boolean, default=False, index=True)
    error = db.Column(db.Text)
    current_file = db.Column(db.String(255))
    lote_id = db.Column(db.String(36), index=True)
    total_registros = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            "job_id": self.job_id,
            "usuario": self.usuario,
            "ente_nombre": self.ente_nombre,
            "archivos": self.archivos or [],
            "progress": self.progress or 0,
            "message": self.message or "",
            "done": bool(self.done),
            "error": self.error,
            "current_file": self.current_file,
            "lote_id": self.lote_id,
            "total_registros": self.total_registros or 0,
        }


class Usuario(db.Model):
    __tablename__ = "usuarios"

    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    nombre_completo = db.Column(db.String(200))
    email = db.Column(db.String(200))
    password_hash = db.Column(db.String(255))
    rol = db.Column(db.String(50), default="auditor")
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)
    ultimo_acceso = db.Column(db.DateTime)

    def to_dict(self):
        return {
            "id": self.id,
            "username": self.username,
            "nombre_completo": self.nombre_completo,
            "email": self.email,
            "rol": self.rol,
            "activo": self.activo,
        }


class Ente(db.Model):
    __tablename__ = 'entes'
    id = db.Column(db.Integer, primary_key=True)
    clave = db.Column(db.String(20), unique=True, nullable=False, index=True)
    codigo = db.Column(db.String(50), nullable=False)
    dd = db.Column(db.String(10))
    dd_match = db.Column(db.String(10), index=True)
    nombre = db.Column(db.String(255), nullable=False)
    siglas = db.Column(db.String(50))
    tipo = db.Column(db.String(100))
    ambito = db.Column(db.String(50))
    activo = db.Column(db.Boolean, default=True)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'clave': self.clave,
            'codigo': self.codigo,
            'dd': self.dd,
            'dd_match': self.dd_match,
            'nombre': self.nombre,
            'siglas': self.siglas,
            'tipo': self.tipo,
            'ambito': self.ambito,
            'activo': self.activo,
            'fecha_registro': self.fecha_registro.isoformat() if self.fecha_registro else None
        }


class ReporteGenerado(db.Model):
    __tablename__ = 'reportes_generados'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    fecha_generacion = db.Column(db.DateTime, default=datetime.utcnow)
    tipo_reporte = db.Column(db.String(50))
    filtros_aplicados = db.Column(db.JSON)
    total_registros = db.Column(db.Integer)
    nombre_archivo = db.Column(db.String(255))
    usuario = db.relationship('Usuario', backref='reportes')

    def to_dict(self):
        return {
            'id': self.id,
            'usuario_id': self.usuario_id,
            'fecha_generacion': self.fecha_generacion.isoformat() if self.fecha_generacion else None,
            'tipo_reporte': self.tipo_reporte,
            'filtros_aplicados': self.filtros_aplicados,
            'total_registros': self.total_registros,
            'nombre_archivo': self.nombre_archivo
        }


def seed_entes_from_catalog():
    """Populate the Ente table from ENTES_CATALOG if it is empty."""
    if Ente.query.first() is not None:
        return
    for entry in ENTES_CATALOG:
        ente = Ente(
            clave=entry["codigo"],
            codigo=entry["codigo"],
            nombre=entry["nombre"],
            siglas=entry.get("siglas", ""),
            tipo=entry.get("tipo", ""),
            ambito=entry.get("ambito", "ESTATAL"),
            activo=True,
        )
        db.session.add(ente)
    try:
        db.session.commit()
        logger.info(f"Seeded {len(ENTES_CATALOG)} entes into DB")
    except Exception:
        db.session.rollback()
        logger.warning("Entes already seeded or seed error, skipping")


# ==================== UTILIDADES ====================

def _norm(s):
    s = str(s or "").strip().lower()
    rep = {"á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ü": "u", "ñ": "n"}
    for k, v in rep.items():
        s = s.replace(k, v)
    s = re.sub(r"\s+", " ", s)
    return s


def _parse_amount(s):
    s = str(s or "").strip()
    if not s:
        return 0.0
    s = s.replace(",", "").replace(" ", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _parse_fecha(s, fallback_mes=1, fallback_ano=None):
    s = str(s or "").strip()
    if not s:
        return None
    s_upper = s.upper()
    if s_upper.startswith("INICIAL") or s_upper == "INICIAL":
        if fallback_ano:
            return date(fallback_ano, fallback_mes, 1)
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _ente_code_from_filename(filename):
    m = re.search(r"_(\d+)\.txt$", filename, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def _split_cuenta_segmentos(cuenta_str):
    parts = [p.strip() for p in str(cuenta_str or "").split(".") if p.strip()]
    keys = ["seg1", "seg2", "seg3", "seg4", "seg5", "seg6"]
    result = {k: "" for k in keys}
    for i, key in enumerate(keys):
        if i < len(parts):
            result[key] = parts[i]
    return result


def _split_cuenta_contable_vertical(cuenta_str):
    """Divide la cuenta contable en componentes verticales (formato SIIF)."""
    s = str(cuenta_str or "").strip().upper()
    s = re.sub(r"[^0-9A-Z]", "", s).ljust(21, "0")
    return {
        "genero": s[0],
        "grupo": s[1],
        "rubro": s[2],
        "cuenta_nivel": s[3],
        "subcuenta": s[4],
        "dependencia": s[5:7],
        "unidad_responsable": s[7:9],
        "centro_costo": s[9:11],
        "proyecto_presupuestario": s[11:13],
        "fuente": s[13],
        "subfuente": s[14:16],
        "tipo_recurso": s[16],
        "partida_presupuestal": s[17:21],
    }


def _hash_transaccion(row):
    parts = [
        _norm(row.get("archivo_origen")),
        _norm(row.get("_source_row_number")),
        _norm(row.get("cuenta_contable")),
        _norm(row.get("nombre_cuenta")),
        _norm(row.get("poliza")),
        _norm(row.get("beneficiario")),
        _norm(row.get("descripcion")),
        _norm(row.get("cheque_folio")),
        _norm(row.get("ente_codigo")),
        str(float(_parse_amount(row.get("cargos")))),
        str(float(_parse_amount(row.get("abonos")))),
    ]
    fecha = row.get("fecha")
    if fecha and not pd.isna(fecha):
        try:
            parts.append(str(fecha))
        except Exception:
            parts.append("")
    else:
        parts.append("")
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _build_balance_summary(df, filename):
    if df is None or df.empty:
        raise ValueError(f"El archivo {filename} no contiene transacciones válidas.")

    total_cargos = float(pd.to_numeric(df["cargos"], errors="coerce").fillna(0.0).sum())
    total_abonos = float(pd.to_numeric(df["abonos"], errors="coerce").fillna(0.0).sum())
    diferencia = total_cargos - total_abonos

    return {
        "filename": filename,
        "total_registros": int(len(df)),
        "total_cargos": total_cargos,
        "total_abonos": total_abonos,
        "diferencia": diferencia,
        "coincide": abs(diferencia) < BALANCE_TOLERANCE,
    }


def _ensure_balanced_file(df, filename):
    summary = _build_balance_summary(df, filename)
    if not summary["coincide"]:
        raise ValueError(
            "Cargo y Abono no coinciden en "
            f"{filename}: cargos={summary['total_cargos']:.2f}, "
            f"abonos={summary['total_abonos']:.2f}, "
            f"diferencia={summary['diferencia']:.2f}"
        )
    return summary


def validate_txt_file_balance(file_data, periodo_mes=1, periodo_ano=None):
    df, filename = _read_one_txt(file_data, periodo_mes, periodo_ano)
    return _ensure_balanced_file(df, filename)


# ==================== PARSER TXT ====================

def _read_one_txt(file_data, periodo_mes=1, periodo_ano=None):
    """
    Lee un auxiliar contable en formato TXT (tabulado) de SAACG4.
    Retorna (DataFrame, filename).
    """
    filename, file_content = file_data
    logger.info(f"Iniciando lectura de archivo TXT: {filename}")
    file_content.seek(0)

    # Leer con encoding latin-1 (los TXT del sistema usan CP1252/ISO-8859-1)
    try:
        raw_bytes = file_content.read()
        try:
            content = raw_bytes.decode("latin-1")
        except Exception:
            content = raw_bytes.decode("utf-8", errors="replace")
    except Exception as e:
        logger.error(f"Error leyendo {filename}: {e}")
        return pd.DataFrame(), filename

    lines = content.splitlines()

    if len(lines) < 3:
        logger.warning(f"Archivo {filename} demasiado corto ({len(lines)} líneas)")
        return pd.DataFrame(), filename

    # Extraer mes del filename si disponible: R1-L1-M12_101.TXT → mes=12
    m_mes = re.search(r"-M(\d+)_", filename, re.IGNORECASE)
    if m_mes:
        periodo_mes = int(m_mes.group(1))
    if periodo_ano is None:
        periodo_ano = datetime.now().year

    # Detectar ente desde el nombre del archivo
    ente_codigo = _ente_code_from_filename(filename)
    ente_info = ENTES_BY_CODIGO.get(ente_codigo, {})

    records = []
    current_cuenta = None
    current_nombre_parts = []

    # Línea 0 = headers, línea 1 = separadores, datos desde línea 2
    for line_idx, line in enumerate(lines):
        if line_idx < 2:
            continue

        cols = line.split("\t")
        cols = [c.strip() for c in cols]

        # Asegurar mínimo de columnas (TXT tiene ~51 columnas)
        while len(cols) < 51:
            cols.append("")

        col0 = cols[0]   # Cuenta
        col1 = cols[1]   # Descripción / summary
        col2 = cols[2]   # Cargos
        col3 = cols[3]   # Abonos
        col4 = cols[4]   # Saldo
        col5 = cols[5]   # Fecha
        col6 = cols[6]   # Póliza
        col7 = cols[7]   # Primer concepto
        col8 = cols[8]   # Segundo concepto
        col9 = cols[9]   # Tercer concepto
        col10 = cols[10]  # Concepto específico
        col11 = cols[11]  # Concepto
        col12 = cols[12]  # Requisición
        col13 = cols[13]  # Cheque/clave
        col14 = cols[14]  # Benef. (código)
        col15 = cols[15]  # Nombre del beneficiario
        col16 = cols[16]  # Obra/acción
        col17 = cols[17]  # Nombre de la obra
        col18 = cols[18]  # Concept (código)
        col19 = cols[19]  # Nombre del concepto
        col20 = cols[20]  # F.F.
        col21 = cols[21]  # Rem.
        col22 = cols[22]  # Nombre de la fuente de financiamiento
        col23 = cols[23]  # Unid
        col24 = cols[24]  # Nombre de la unidad administrativa
        col25 = cols[25]  # Programa
        col26 = cols[26]  # Nombre del programa
        col27 = cols[27]  # Co
        col28 = cols[28]  # Nombre del componente
        col29 = cols[29]  # Ac
        col30 = cols[30]  # Nombre de la actividad
        col31 = cols[31]  # Func
        col32 = cols[32]  # Nombre de la función
        col33 = cols[33]  # Contrato
        col34 = cols[34]  # Ficha
        col35 = cols[35]  # Referencia
        col36 = cols[36]  # Persona
        col37 = cols[37]  # RFC
        col38 = cols[38]  # Nombre
        col39 = cols[39]  # Número de bien
        col40 = cols[40]  # Nombre del bien
        col41 = cols[41]  # Unidades
        col42 = cols[42]  # Partida presupuestal
        col43 = cols[43]  # Nombre de la partida presupuestal
        col44 = cols[44]  # Rg
        col45 = cols[45]  # Nombre de la región
        col46 = cols[46]  # Mun
        col47 = cols[47]  # Nombre del municipio
        col48 = cols[48]  # Cargos f
        col49 = cols[49]  # Abonos f
        col50 = cols[50]  # Saldo F

        # Detectar encabezado de cuenta contable (col0 tiene el código)
        if col0 and re.match(r"^\d+\.\d+", col0):
            current_cuenta = col0
            current_nombre_parts = [col1] if col1 else []
            continue

        if not current_cuenta:
            continue

        # Saltar líneas completamente vacías
        if not col0 and not col1 and not col2 and not col3:
            continue

        # Detectar líneas de resumen (Saldo inicial, Suma periodo, Saldo final)
        col1_clean = col1.rstrip(".").strip().lower()
        if (
            col1_clean.startswith("saldo inicial")
            or col1_clean.startswith("suma periodo")
            or col1_clean.startswith("saldo final")
        ):
            continue

        # Detectar líneas de descripción (col0 vacío, col1 no vacío, sin montos)
        if not col0 and col1 and not col2 and not col3:
            if col1 not in current_nombre_parts:
                current_nombre_parts.append(col1)
            continue

        # Detectar fila de transacción (hay monto en cargos o abonos)
        if not col2 and not col3:
            continue

        fecha = _parse_fecha(col5, periodo_mes, periodo_ano)
        if not fecha:
            continue

        # Expandir descripciones (hasta 3 líneas adicionales)
        descs = (current_nombre_parts + [""] * 4)[:4]
        nombre_cuenta = descs[0]

        # Preferir col7 (primer concepto) como descripción; si está vacío, usar col1
        descripcion = col7 or col1

        record = {
            "cuenta_contable": current_cuenta,
            "nombre_cuenta": nombre_cuenta,
            "descripcion_1": descs[1],
            "descripcion_2": descs[2],
            "descripcion_3": descs[3],
            "fecha": fecha.strftime("%Y-%m-%d"),
            "poliza": col6,
            "primer_concepto": col7,
            "segundo_concepto": col8,
            "tercer_concepto": col9,
            "concepto_especifico": col10,
            "concepto": col11,
            "requisicion": col12,
            "cheque_folio": col13,
            "beneficiario_codigo": col14,
            "beneficiario": col15,
            "descripcion": descripcion,
            "obra_accion": col16,
            "nombre_obra": col17,
            "concept_codigo": col18,
            "nombre_concepto": col19,
            "ff": col20,
            "rem": col21,
            "nombre_fuente_financiamiento": col22,
            "unidad_codigo": col23,
            "nombre_unidad_administrativa": col24,
            "programa": col25,
            "nombre_programa": col26,
            "componente": col27,
            "nombre_componente": col28,
            "actividad": col29,
            "nombre_actividad": col30,
            "funcion": col31,
            "nombre_funcion": col32,
            "contrato": col33,
            "ficha": col34,
            "referencia": col35,
            "persona": col36,
            "rfc": col37,
            "nombre_persona": col38,
            "numero_bien": col39,
            "nombre_bien": col40,
            "unidades": col41,
            "partida_presupuestal_txt": col42,
            "nombre_partida_presupuestal": col43,
            "region": col44,
            "nombre_region": col45,
            "municipio": col46,
            "nombre_municipio": col47,
            "cargos": _parse_amount(col2),
            "abonos": _parse_amount(col3),
            "saldo_final": _parse_amount(col4),
            "cargos_f": _parse_amount(col48),
            "abonos_f": _parse_amount(col49),
            "saldo_f": _parse_amount(col50),
            "ente_codigo": ente_codigo or "",
            "ente_nombre": ente_info.get("nombre", ""),
            "ente_siglas": ente_info.get("siglas", ""),
            "_source_row_number": line_idx + 1,
        }
        records.append(record)

    if not records:
        logger.warning(f"No se encontraron transacciones en {filename}")
        return pd.DataFrame(), filename

    df = pd.DataFrame(records)
    logger.info(f"✓ {len(df)} transacciones extraídas de {filename}")
    return df, filename


# ==================== PROCESAMIENTO ====================

def process_files_to_database(
    file_list: List[Tuple[str, BytesIO]],
    usuario: str = "sistema",
    progress_callback: Optional[Callable] = None,
    periodo_ano: Optional[int] = None,
):
    """
    Procesa archivos TXT y guarda en base de datos.
    Retorna (lote_id, total_registros).
    """
    def report(p, m, current_file=None):
        if progress_callback:
            progress_callback(p, m, current_file)
        else:
            print(f"[{p}%] {m}")

    lote_id = str(uuid.uuid4())

    lote = LoteCarga(
        lote_id=lote_id,
        usuario=usuario,
        archivos=[f[0] for f in file_list],
        estado="procesando",
    )
    db.session.add(lote)
    db.session.commit()

    try:
        report(5, f"Leyendo {len(file_list)} archivo(s)...")
        frames = []
        archivos_procesados = []
        archivos_fallidos = []
        total_files = len(file_list)
        completed = 0

        with ThreadPoolExecutor(max_workers=min(4, len(file_list))) as ex:
            futures = {
                ex.submit(_read_one_txt, f, 1, periodo_ano): f
                for f in file_list
            }
            for fut in as_completed(futures):
                completed += 1
                pct = 5 + int((completed / total_files) * 25)
                try:
                    df, fname = fut.result()
                    if not df.empty:
                        _ensure_balanced_file(df, fname)
                        df["archivo_origen"] = fname
                        frames.append(df)
                        archivos_procesados.append(fname)
                        report(pct, f"Leído: {fname}", fname)
                    else:
                        archivos_fallidos.append(futures[fut][0])
                        report(pct, f"Sin datos: {futures[fut][0]}", futures[fut][0])
                except Exception as e:
                    fname = futures[fut][0]
                    logger.error(f"Error en {fname}: {e}\n{traceback.format_exc()}")
                    archivos_fallidos.append(fname)
                    report(pct, f"Error en: {fname}", fname)

        if not frames:
            msg = f"No se procesó ningún archivo. Fallidos: {', '.join(archivos_fallidos)}"
            lote.estado = "error"
            lote.mensaje = msg
            db.session.commit()
            raise ValueError(msg)

        report(32, "Combinando datos...")
        base = pd.concat(frames, ignore_index=True)

        # Limpiar campos de texto
        text_cols = [
            "poliza", "beneficiario", "cheque_folio", "descripcion",
            "descripcion_1", "descripcion_2", "descripcion_3",
            "primer_concepto", "segundo_concepto", "tercer_concepto",
            "concepto_especifico", "concepto", "requisicion",
            "beneficiario_codigo", "obra_accion", "nombre_obra",
            "concept_codigo", "nombre_concepto", "ff", "rem",
            "nombre_fuente_financiamiento", "unidad_codigo",
            "nombre_unidad_administrativa", "programa", "nombre_programa",
            "componente", "nombre_componente", "actividad", "nombre_actividad",
            "funcion", "nombre_funcion", "contrato", "ficha", "referencia",
            "persona", "rfc", "nombre_persona", "numero_bien", "nombre_bien",
            "unidades", "partida_presupuestal_txt", "nombre_partida_presupuestal",
            "region", "nombre_region", "municipio", "nombre_municipio",
        ]
        for col in text_cols:
            if col in base.columns:
                base[col] = base[col].fillna("").astype(str).str.strip()

        # Dividir cuenta en segmentos
        report(40, "Procesando cuentas contables...")
        segs = base["cuenta_contable"].apply(_split_cuenta_segmentos)
        for key in ("seg1", "seg2", "seg3", "seg4", "seg5", "seg6"):
            base[key] = segs.apply(lambda d: d.get(key, ""))

        # Split into vertical components (SIIF format)
        report(45, "Procesando componentes verticales...")
        componentes_vertical = base["cuenta_contable"].apply(_split_cuenta_contable_vertical)
        for key in ["genero", "grupo", "rubro", "cuenta_nivel", "subcuenta", "dependencia",
                     "unidad_responsable", "centro_costo", "proyecto_presupuestario",
                     "fuente", "subfuente", "tipo_recurso", "partida_presupuestal"]:
            base[key] = componentes_vertical.apply(lambda x, k=key: x[k])

        # Convertir montos
        report(50, "Procesando montos...")
        for col in ("cargos", "abonos", "saldo_final", "cargos_f", "abonos_f", "saldo_f"):
            base[col] = pd.to_numeric(base[col], errors="coerce").fillna(0.0)

        # Hash para deduplicación
        report(60, "Calculando hashes...")
        base["hash_registro"] = base.apply(_hash_transaccion, axis=1)

        # Eliminar hashes duplicados dentro del lote
        base = base.drop_duplicates(subset=["hash_registro"], keep="first")

        # Verificar hashes ya existentes en la DB
        report(68, "Verificando duplicados...")
        existing_hashes = {
            h for (h,) in db.session.query(Transaccion.hash_registro).all()
        }
        duplicate_mask = base["hash_registro"].isin(existing_hashes)
        if duplicate_mask.any():
            duplicate_rows = base.loc[duplicate_mask, ["archivo_origen", "hash_registro"]]
            duplicate_counts = (
                duplicate_rows.groupby("archivo_origen")["hash_registro"]
                .count()
                .to_dict()
            )
            file_counts = base.groupby("archivo_origen")["hash_registro"].count().to_dict()
            partial_files = sorted(
                filename
                for filename, count in duplicate_counts.items()
                if count < file_counts.get(filename, 0)
            )
            if partial_files:
                raise ValueError(
                    "Se detectaron registros previamente cargados en "
                    f"{', '.join(partial_files)}. Se rechaza la carga para evitar resultados parciales."
                )
        base = base[~base["hash_registro"].isin(existing_hashes)]

        if base.empty:
            lote.estado = "completado"
            lote.total_registros = 0
            lote.mensaje = "Todos los registros ya existen en la base de datos."
            db.session.commit()
            return lote_id, 0

        # Insertar en la DB por lotes
        report(72, "Guardando registros...")
        chunk_size = 500
        total = len(base)
        inserted = 0

        for start in range(0, total, chunk_size):
            chunk = base.iloc[start : start + chunk_size]
            for _, row in chunk.iterrows():
                fecha_val = row.get("fecha")
                if isinstance(fecha_val, str):
                    try:
                        fecha_obj = datetime.strptime(fecha_val, "%Y-%m-%d").date()
                    except ValueError:
                        continue
                elif isinstance(fecha_val, (date, datetime)):
                    fecha_obj = fecha_val if isinstance(fecha_val, date) else fecha_val.date()
                else:
                    continue

                def _s(key, default=""):
                    v = row.get(key, default)
                    return str(v) if v is not None and str(v) != "nan" else default

                def _f(key):
                    try:
                        v = row.get(key, 0)
                        return float(v) if v is not None and str(v) != "nan" else 0.0
                    except (ValueError, TypeError):
                        return 0.0

                t = Transaccion(
                    lote_id=lote_id,
                    archivo_origen=_s("archivo_origen"),
                    usuario_carga=usuario,
                    cuenta_contable=_s("cuenta_contable"),
                    nombre_cuenta=_s("nombre_cuenta"),
                    descripcion_1=_s("descripcion_1"),
                    descripcion_2=_s("descripcion_2"),
                    descripcion_3=_s("descripcion_3"),
                    seg1=_s("seg1"),
                    seg2=_s("seg2"),
                    seg3=_s("seg3"),
                    seg4=_s("seg4"),
                    seg5=_s("seg5"),
                    seg6=_s("seg6"),
                    genero=_s("genero"),
                    grupo=_s("grupo"),
                    rubro=_s("rubro"),
                    cuenta_nivel=_s("cuenta_nivel"),
                    subcuenta=_s("subcuenta"),
                    dependencia=_s("dependencia"),
                    unidad_responsable=_s("unidad_responsable"),
                    centro_costo=_s("centro_costo"),
                    proyecto_presupuestario=_s("proyecto_presupuestario"),
                    fuente=_s("fuente"),
                    subfuente=_s("subfuente"),
                    tipo_recurso=_s("tipo_recurso"),
                    partida_presupuestal=_s("partida_presupuestal"),
                    ente_codigo=_s("ente_codigo"),
                    ente_nombre=_s("ente_nombre"),
                    ente_siglas=_s("ente_siglas"),
                    fecha_transaccion=fecha_obj,
                    poliza=_s("poliza"),
                    primer_concepto=_s("primer_concepto"),
                    segundo_concepto=_s("segundo_concepto"),
                    tercer_concepto=_s("tercer_concepto"),
                    concepto_especifico=_s("concepto_especifico"),
                    concepto=_s("concepto"),
                    requisicion=_s("requisicion"),
                    cheque_folio=_s("cheque_folio"),
                    beneficiario_codigo=_s("beneficiario_codigo"),
                    beneficiario=_s("beneficiario"),
                    descripcion=_s("descripcion"),
                    obra_accion=_s("obra_accion"),
                    nombre_obra=_s("nombre_obra"),
                    concept_codigo=_s("concept_codigo"),
                    nombre_concepto=_s("nombre_concepto"),
                    ff=_s("ff"),
                    rem=_s("rem"),
                    nombre_fuente_financiamiento=_s("nombre_fuente_financiamiento"),
                    unidad_codigo=_s("unidad_codigo"),
                    nombre_unidad_administrativa=_s("nombre_unidad_administrativa"),
                    programa=_s("programa"),
                    nombre_programa=_s("nombre_programa"),
                    componente=_s("componente"),
                    nombre_componente=_s("nombre_componente"),
                    actividad=_s("actividad"),
                    nombre_actividad=_s("nombre_actividad"),
                    funcion=_s("funcion"),
                    nombre_funcion=_s("nombre_funcion"),
                    contrato=_s("contrato"),
                    ficha=_s("ficha"),
                    referencia=_s("referencia"),
                    persona=_s("persona"),
                    rfc=_s("rfc"),
                    nombre_persona=_s("nombre_persona"),
                    numero_bien=_s("numero_bien"),
                    nombre_bien=_s("nombre_bien"),
                    unidades=_s("unidades"),
                    partida_presupuestal_txt=_s("partida_presupuestal_txt"),
                    nombre_partida_presupuestal=_s("nombre_partida_presupuestal"),
                    region=_s("region"),
                    nombre_region=_s("nombre_region"),
                    municipio=_s("municipio"),
                    nombre_municipio=_s("nombre_municipio"),
                    cargos=_f("cargos"),
                    abonos=_f("abonos"),
                    saldo_final=_f("saldo_final"),
                    cargos_f=_f("cargos_f"),
                    abonos_f=_f("abonos_f"),
                    saldo_f=_f("saldo_f"),
                    hash_registro=_s("hash_registro"),
                )
                db.session.add(t)

            db.session.commit()
            inserted += len(chunk)
            pct = 72 + int((inserted / total) * 25)
            report(min(pct, 97), f"Guardados {inserted:,}/{total:,} registros...")

        lote.estado = "completado"
        lote.total_registros = inserted
        lote.mensaje = f"{inserted:,} registros procesados."
        db.session.commit()

        report(100, f"Completado. {inserted:,} registros guardados.")
        logger.info(f"✓ Lote {lote_id}: {inserted} registros de {archivos_procesados}")
        return lote_id, inserted

    except Exception as e:
        db.session.rollback()
        lote.estado = "error"
        lote.mensaje = str(e)
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()
        logger.error(f"Error en process_files_to_database: {e}\n{traceback.format_exc()}")
        raise


def _decode_txt_bytes(raw_bytes):
    try:
        return raw_bytes.decode("latin-1")
    except Exception:
        return raw_bytes.decode("utf-8", errors="replace")


def _coerce_example_numeric(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value

    text = str(value).replace(",", "").strip()
    if not text or set(text) == {"-"}:
        return value

    try:
        number = float(text)
    except ValueError:
        return value

    return int(number) if number.is_integer() else number


def _parse_txt_for_example_export(file_data):
    filename, file_content = file_data
    file_content.seek(0)
    raw_bytes = file_content.read()
    content = _decode_txt_bytes(raw_bytes)
    lines = content.splitlines()

    if len(lines) < 2:
        raise ValueError(f"El archivo {filename} no contiene encabezados suficientes")

    source_headers = [str(value or "").strip() for value in lines[0].split("\t")]
    separator_values = [str(value or "").strip() for value in lines[1].split("\t")]

    while len(separator_values) < len(source_headers):
        separator_values.append("")

    expanded_headers = (
        source_headers[:2]
        + ["Descripción 1", "Descripción 2", "Descripción 3"]
        + source_headers[2:]
    )

    description_limit = 4
    current_cuenta = ""
    current_descriptions = []
    records = []

    for raw_line in lines[2:]:
        cols = [str(value or "").strip() for value in raw_line.split("\t")]
        while len(cols) < len(source_headers):
            cols.append("")
        cols = cols[:len(source_headers)]

        if not any(cols):
            continue

        cuenta = cols[0]
        descripcion = cols[1]

        if cuenta and re.match(r"^\d+(?:\.\d+)+$", cuenta):
            current_cuenta = cuenta
            current_descriptions = [descripcion] if descripcion else []
            continue

        if not current_cuenta:
            continue

        descripcion_normalizada = descripcion.rstrip(".").strip().lower()
        if (
            descripcion_normalizada.startswith("saldo inicial")
            or descripcion_normalizada.startswith("suma periodo")
            or descripcion_normalizada.startswith("saldo final")
        ):
            continue

        has_transaction_payload = any(
            cols[index]
            for index in (2, 3, 4, 5, 6)
            if index < len(cols)
        )

        if not cuenta and descripcion and not has_transaction_payload:
            if descripcion not in current_descriptions and len(current_descriptions) < description_limit:
                current_descriptions.append(descripcion)
            continue

        if not has_transaction_payload:
            continue

        descriptions = (current_descriptions + [""] * description_limit)[:description_limit]
        row = {
            source_headers[0]: current_cuenta,
            source_headers[1]: descriptions[0],
            "Descripción 1": descriptions[1],
            "Descripción 2": descriptions[2],
            "Descripción 3": descriptions[3],
        }

        for header, value in zip(source_headers[2:], cols[2:]):
            row[header] = value

        records.append(row)

    separator_row = {header: "" for header in expanded_headers}
    separator_row[source_headers[0]] = separator_values[0] if separator_values else ""
    if len(separator_values) > 1:
        separator_row[source_headers[1]] = separator_values[1]
    for header, value in zip(source_headers[2:], separator_values[2:]):
        separator_row[header] = value

    return expanded_headers, separator_row, records


def build_example_output_workbook(input_dir=None, output_path=None, periodo_ano=None):
    input_dir = Path(input_dir or Path.cwd() / "example_SCG4" / "input")
    output_path = Path(output_path or Path.cwd() / "example_SCG4" / "output" / "output.xlsx")

    txt_files = sorted(
        list(input_dir.glob("*.TXT")) + list(input_dir.glob("*.txt"))
    )
    if not txt_files:
        raise FileNotFoundError(f"No se encontraron archivos TXT en {input_dir}")

    all_rows = []
    headers = None
    separator_row = None

    for path in txt_files:
        raw_bytes = path.read_bytes()
        validate_txt_file_balance(
            (path.name, BytesIO(raw_bytes)),
            periodo_ano=periodo_ano,
        )
        file_headers, file_separator, file_rows = _parse_txt_for_example_export(
            (path.name, BytesIO(raw_bytes))
        )

        if headers is None:
            headers = file_headers
            separator_row = file_separator

        all_rows.extend(file_rows)

    if headers is None:
        raise ValueError("No se pudieron construir encabezados para la exportación")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows_to_export = [separator_row] + all_rows
    df = pd.DataFrame(rows_to_export, columns=headers)

    for column in ("Cargos", "Abonos", "Saldo", "Cargos f", "Abonos f", "Saldo F"):
        if column in df.columns:
            df[column] = df[column].apply(_coerce_example_numeric)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Procesado")

    workbook = load_workbook(output_path)
    sheet = workbook["Procesado"]
    for column in ("Cargos", "Abonos", "Saldo", "Cargos f", "Abonos f", "Saldo F"):
        if column not in headers:
            continue
        column_index = headers.index(column) + 1
        for row_index in range(3, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = _coerce_example_numeric(cell.value)
    workbook.save(output_path)

    return {
        "output_path": str(output_path),
        "files_processed": [path.name for path in txt_files],
        "total_rows": int(len(df)),
    }


def ensure_transacciones_schema():
    inspector = inspect(db.engine)
    if "transacciones" not in inspector.get_table_names():
        return

    existing_columns = {
        column["name"]
        for column in inspector.get_columns("transacciones")
    }
    expected_columns = {
        "genero": "VARCHAR(1)",
        "grupo": "VARCHAR(1)",
        "rubro": "VARCHAR(1)",
        "cuenta_nivel": "VARCHAR(1)",
        "subcuenta": "VARCHAR(1)",
        "dependencia": "VARCHAR(2)",
        "unidad_responsable": "VARCHAR(2)",
        "centro_costo": "VARCHAR(2)",
        "proyecto_presupuestario": "VARCHAR(2)",
        "fuente": "VARCHAR(1)",
        "subfuente": "VARCHAR(2)",
        "tipo_recurso": "VARCHAR(1)",
        "partida_presupuestal": "VARCHAR(4)",
    }

    missing_columns = [
        (name, column_type)
        for name, column_type in expected_columns.items()
        if name not in existing_columns
    ]
    if not missing_columns:
        return

    for column_name, column_type in missing_columns:
        db.session.execute(
            text(
                f"ALTER TABLE transacciones "
                f"ADD COLUMN {column_name} {column_type}"
            )
        )
    db.session.commit()
